"""Append one row per analyzed chunk (image region) to an Excel workbook,
one sheet per source image, with embedded plot thumbnails.

Schema is intentionally flat (one row = one chunk) so it stays a plain
sortable/filterable table in Excel as more chunks of the same image, or
entirely new images, get analyzed over time -- appending never requires a
schema change, only calling append_chunk_to_workbook() again. `region.py`'s
crop-selection logic is untouched by this module; whatever region info is
available (or None, for a --full-image run) is simply recorded.

The four aggregate helpers below (tortuosity_stats, curvature_stats,
corner_agreement_counts, anisotropy_classification) are the single source
of truth for numbers that used to be computed inline, only in
analyze_image.py's print block -- analyze_image.py now imports them too,
so the console report and the spreadsheet row can never silently disagree.
"""

from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

from .anisotropy import AnisotropyResult
from .binarize import BinarizeResult
from .corners import CornerCrossCheck
from .curvature import CurvatureScanResult
from .graph import GraphResult
from .junctions import JunctionAnalysisResult
from .kinks import KinkScanResult
from .region import Region
from .skeleton import SkeletonResult
from .summary_plots import render_curvature_histogram, render_orientation_rose

MAX_SHEET_NAME_LEN = 31
INVALID_SHEET_CHARS = "[]:*?/\\"

THUMBNAIL_WIDTH_PX = 220
ROW_HEIGHT_FOR_THUMBNAILS = 130

ANISOTROPY_GRID_LIKE_THRESHOLD = 0.5
ANISOTROPY_MUDCRACK_LIKE_THRESHOLD = 0.15
# [PLACEHOLDER] not calibrated against a real-image junction-type census;
# same thresholds already used for the console report's qualitative read.

COLUMN_HEADERS: list[str] = [
    # run identity
    "timestamp_iso",
    "image_filename",
    "image_stem",
    "region_description",
    "region_row_start_px",
    "region_row_stop_px",
    "region_col_start_px",
    "region_col_stop_px",
    "region_corner_frac",
    "region_edge_margin_frac",
    # stage 1: binarize
    "otsu_threshold_0to255",
    "foreground_fraction_0to1",
    "binarize_sanity_ok",
    "binarize_sanity_band_lo",
    "binarize_sanity_band_hi",
    # stage 2: skeleton
    "skeleton_px_pre_prune",
    "skeleton_px_post_prune",
    "spur_px_threshold_PLACEHOLDER",
    "n_spurs_pruned",
    "n_fragments_pruned",
    "prune_iters_used",
    # stage 3: graph
    "n_nodes_total",
    "n_endpoints_deg1",
    "n_junctions_deg3",
    "n_junctions_deg_ge4",
    "n_edges",
    # curvature/tortuosity
    "curvature_window_px_PLACEHOLDER",
    "tortuosity_mean_arc_over_chord",
    "tortuosity_median_arc_over_chord",
    "mean_curvature_px_inv",
    "max_curvature_px_inv",
    "n_edges_scanned_curvature",
    "n_edges_no_curvature_profile",
    # anisotropy
    "anisotropy_window_px_PLACEHOLDER",
    "anisotropy_index_0to1",
    "dominant_bearing_deg",
    "anisotropy_n_samples",
    "anisotropy_total_weighted_length_px",
    "anisotropy_qualitative_class",
    # junctions
    "junction_annulus_inner_px_PLACEHOLDER",
    "junction_annulus_outer_px_PLACEHOLDER",
    "junction_y_angle_tol_deg_PLACEHOLDER",
    "junction_t_straight_tol_deg_PLACEHOLDER",
    "junction_t_right_tol_deg_PLACEHOLDER",
    "n_deg3_junctions_total",
    "n_junctions_T",
    "n_junctions_Y",
    "n_junctions_ambiguous",
    "n_junctions_insufficient_data",
    "n_junctions_deg_ge4_unclassified",
    # kinks
    "kink_window_px_PLACEHOLDER",
    "kink_min_turn_deg_PLACEHOLDER",
    "n_kinks_flagged",
    "n_edges_scanned_kinks",
    # corner cross-check
    "n_corner_checks_total",
    "n_corner_agree",
    "n_corner_disagree",
    "n_corner_unresolved",
]

IMAGE_COLUMN_HEADERS: list[str] = [
    "orientation_rose_plot",
    "curvature_histogram_plot",
    "skeleton_overlay_plot",
]


def tortuosity_stats(curvature_result: CurvatureScanResult) -> tuple[float | None, float | None]:
    """(mean, median) tortuosity across edges with a non-None value."""
    values = [e.tortuosity for e in curvature_result.edges if e.tortuosity is not None]
    if not values:
        return None, None
    return float(np.mean(values)), float(np.median(values))


def curvature_stats(curvature_result: CurvatureScanResult) -> tuple[float | None, float | None]:
    """(mean, max) of per-edge mean|curvature| across edges with a profile."""
    values = [
        e.mean_abs_curvature_px_inv for e in curvature_result.edges if e.mean_abs_curvature_px_inv is not None
    ]
    if not values:
        return None, None
    return float(np.mean(values)), float(np.max(values))


def corner_agreement_counts(corner_cross_check: list[CornerCrossCheck]) -> tuple[int, int, int]:
    """(n_agree, n_disagree, n_unresolved) across the corner cross-check list."""
    n_agree = sum(cc.agrees_with_tangent_fit is True for cc in corner_cross_check)
    n_disagree = sum(cc.agrees_with_tangent_fit is False for cc in corner_cross_check)
    n_unresolved = sum(cc.label is None for cc in corner_cross_check)
    return n_agree, n_disagree, n_unresolved


def anisotropy_classification(anisotropy_result: AnisotropyResult) -> str:
    """[interpreted] qualitative read of the anisotropy index -- read
    alongside the orientation histogram, not in place of it (see
    anisotropy.py's documented 2nd-order-tensor blind spot)."""
    if anisotropy_result.anisotropy_index >= ANISOTROPY_GRID_LIKE_THRESHOLD:
        return "grid-like / rectilinear"
    elif anisotropy_result.anisotropy_index < ANISOTROPY_MUDCRACK_LIKE_THRESHOLD:
        return "mudcrack-like / isotropic"
    else:
        return "intermediate"


def _base_sheet_name(stem: str) -> str:
    cleaned = "".join(c for c in stem if c not in INVALID_SHEET_CHARS).strip() or "sheet"
    return cleaned[:MAX_SHEET_NAME_LEN]


def _sanitize_sheet_name(stem: str, existing_names: set[str]) -> str:
    """Strip invalid Excel sheet-name characters, truncate to 31 chars,
    and disambiguate with a numeric suffix if that collides with an
    already-used name."""
    base = _base_sheet_name(stem)
    if base not in existing_names:
        return base
    suffix_num = 2
    while True:
        suffix = f"-{suffix_num}"
        candidate = base[: MAX_SHEET_NAME_LEN - len(suffix)] + suffix
        if candidate not in existing_names:
            return candidate
        suffix_num += 1


def _get_or_create_sheet(wb: openpyxl.Workbook, image_stem: str) -> tuple[Worksheet, bool]:
    """Reuse the existing sheet for this exact image (more chunks of the
    same image), or create a new one -- disambiguating only when a
    *different* image's stem happens to collide after sanitization
    /truncation. Distinguished by checking the first data row's
    image_stem column, not just the sheet name, since sanitization can
    lose information (stripped characters, truncation) that two distinct
    stems could otherwise share.

    Returns (worksheet, is_new_sheet).
    """
    base = _base_sheet_name(image_stem)
    image_stem_col_idx = COLUMN_HEADERS.index("image_stem") + 1

    if base in wb.sheetnames:
        existing_ws = wb[base]
        if existing_ws.max_row >= 2 and existing_ws.cell(row=2, column=image_stem_col_idx).value == image_stem:
            return existing_ws, False
        sheet_name = _sanitize_sheet_name(image_stem, set(wb.sheetnames))
    else:
        sheet_name = base

    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMN_HEADERS + IMAGE_COLUMN_HEADERS)
    ws.freeze_panes = "A2"
    return ws, True


def _embed_thumbnail(ws: Worksheet, png_path: Path, cell: str) -> None:
    native_w, native_h = PILImage.open(png_path).size
    scale = THUMBNAIL_WIDTH_PX / native_w
    xl_img = XLImage(str(png_path))
    xl_img.width = THUMBNAIL_WIDTH_PX
    xl_img.height = native_h * scale
    ws.add_image(xl_img, cell)


def append_chunk_to_workbook(
    xlsx_path: str | Path,
    *,
    image_path: Path,
    region_desc: str,
    region: Region | None,
    full_image_shape: tuple[int, int] | None,
    binarize_result: BinarizeResult,
    skeleton_result: SkeletonResult,
    graph_result: GraphResult,
    curvature_result: CurvatureScanResult,
    anisotropy_result: AnisotropyResult,
    junction_result: JunctionAnalysisResult,
    kink_result: KinkScanResult,
    corner_cross_check: list[CornerCrossCheck],
    overlay_png_path: Path,
    out_dir: Path,
) -> None:
    """Append one row (this chunk's results) to xlsx_path, creating the
    workbook/sheet/header if they don't exist yet. `region` is None for a
    --full-image run (analyze_image.py's main() never constructs a Region
    in that branch) -- full_image_shape is used instead to fill the
    region_row/col_stop columns.
    """
    xlsx_path = Path(xlsx_path)
    out_dir = Path(out_dir)
    plots_dir = out_dir / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    stem_suffix = "full" if region is None else "corner"
    rose_path = plots_dir / f"{image_path.stem}_{stem_suffix}_rose.png"
    curv_hist_path = plots_dir / f"{image_path.stem}_{stem_suffix}_curvature_hist.png"
    render_orientation_rose(anisotropy_result, rose_path)
    render_curvature_histogram(curvature_result, curv_hist_path)

    wb = openpyxl.load_workbook(xlsx_path) if xlsx_path.exists() else openpyxl.Workbook()

    ws, is_new_sheet = _get_or_create_sheet(wb, image_path.stem)

    if "Sheet" in wb.sheetnames and wb["Sheet"].title != ws.title and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
        wb.remove(wb["Sheet"])

    if region is not None:
        row_start, row_stop = region.row_slice.start, region.row_slice.stop
        col_start, col_stop = region.col_slice.start, region.col_slice.stop
        corner_frac = region.corner_frac
        edge_margin_frac = region.edge_margin_frac
    else:
        full_h, full_w = full_image_shape
        row_start, row_stop = 0, full_h
        col_start, col_stop = 0, full_w
        corner_frac = None
        edge_margin_frac = None

    tortuosity_mean, tortuosity_median = tortuosity_stats(curvature_result)
    curvature_mean, curvature_max = curvature_stats(curvature_result)
    n_agree, n_disagree, n_unresolved = corner_agreement_counts(corner_cross_check)
    classification = anisotropy_classification(anisotropy_result)
    n_edges_no_profile = sum(1 for e in curvature_result.edges if e.curvature_profile_px_inv is None)

    row_values = [
        datetime.now().isoformat(),
        image_path.name,
        image_path.stem,
        region_desc,
        row_start,
        row_stop,
        col_start,
        col_stop,
        corner_frac,
        edge_margin_frac,
        binarize_result.threshold,
        binarize_result.foreground_fraction,
        binarize_result.sanity_ok,
        binarize_result.sanity_band[0],
        binarize_result.sanity_band[1],
        skeleton_result.n_pixels_pre_prune,
        skeleton_result.n_pixels_post_prune,
        skeleton_result.spur_px_threshold,
        skeleton_result.n_spurs_pruned,
        skeleton_result.n_fragments_pruned,
        skeleton_result.iters_used,
        len(graph_result.node_ids),
        graph_result.n_endpoints,
        graph_result.n_junctions_deg3,
        graph_result.n_junctions_deg_ge4,
        graph_result.n_edges,
        curvature_result.window_px,
        tortuosity_mean,
        tortuosity_median,
        curvature_mean,
        curvature_max,
        curvature_result.n_edges_scanned,
        n_edges_no_profile,
        anisotropy_result.window_px,
        anisotropy_result.anisotropy_index,
        anisotropy_result.dominant_bearing_deg,
        anisotropy_result.n_samples,
        anisotropy_result.total_weighted_length_px,
        classification,
        junction_result.inner_radius_px,
        junction_result.outer_radius_px,
        junction_result.y_angle_tol_deg,
        junction_result.t_straight_tol_deg,
        junction_result.t_right_tol_deg,
        junction_result.n_deg3_total,
        junction_result.n_t,
        junction_result.n_y,
        junction_result.n_ambiguous,
        junction_result.n_insufficient_data,
        junction_result.n_deg_ge4_unclassified,
        kink_result.window_px,
        kink_result.min_turn_deg,
        len(kink_result.kinks),
        kink_result.n_edges_scanned,
        len(corner_cross_check),
        n_agree,
        n_disagree,
        n_unresolved,
    ]
    assert len(row_values) == len(COLUMN_HEADERS)

    next_row = ws.max_row + 1
    for col_idx, value in enumerate(row_values, start=1):
        ws.cell(row=next_row, column=col_idx, value=value)

    image_start_col = len(COLUMN_HEADERS) + 1
    for offset, png_path in enumerate([rose_path, curv_hist_path, overlay_png_path]):
        col_letter = get_column_letter(image_start_col + offset)
        _embed_thumbnail(ws, Path(png_path), f"{col_letter}{next_row}")
    ws.row_dimensions[next_row].height = ROW_HEIGHT_FOR_THUMBNAILS

    wb.save(xlsx_path)
