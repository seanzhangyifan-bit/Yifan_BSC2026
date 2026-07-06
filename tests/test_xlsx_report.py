"""Correctness checks for the Excel export: appending more chunks of the
same image adds rows (not overwrite), appending a different image creates
a new sheet (existing sheet untouched), and the sheet-name sanitizer
handles invalid characters / long names / truncation collisions.
"""

from pathlib import Path

import numpy as np
import openpyxl
import pytest

from src.crackgraph.anisotropy import compute_anisotropy
from src.crackgraph.binarize import binarize
from src.crackgraph.corners import cross_check_junctions, find_background_contours
from src.crackgraph.curvature import compute_edge_curvature
from src.crackgraph.graph import extract_graph
from src.crackgraph.junctions import classify_junctions
from src.crackgraph.kinks import find_kinks
from src.crackgraph.overlay import render_overlay
from src.crackgraph.skeleton import skeletonize_and_prune
from src.crackgraph.synthetic import generate_t_junction
from src.crackgraph.xlsx_report import (
    COLUMN_HEADERS,
    IMAGE_COLUMN_HEADERS,
    _sanitize_sheet_name,
    append_chunk_to_workbook,
)


def _run_full_pipeline(gray):
    binarize_result = binarize(gray, sanity_band=(0.0, 1.0))
    skeleton_result = skeletonize_and_prune(binarize_result.mask, source_image=gray, spur_px=3)
    graph_result = extract_graph(skeleton_result.skeleton)
    curvature_result = compute_edge_curvature(skeleton_result.skeleton, graph_result)
    anisotropy_result = compute_anisotropy(skeleton_result.skeleton, graph_result)
    junction_result = classify_junctions(skeleton_result.skeleton, graph_result, medial_radius=skeleton_result.medial_radius)
    kink_result = find_kinks(skeleton_result.skeleton, graph_result)
    background_contours = find_background_contours(skeleton_result.mask_clean)
    corner_cross_check = cross_check_junctions(
        junction_result,
        background_contours,
        y_angle_tol_deg=junction_result.y_angle_tol_deg,
        t_straight_tol_deg=junction_result.t_straight_tol_deg,
        t_right_tol_deg=junction_result.t_right_tol_deg,
    )
    return {
        "binarize_result": binarize_result,
        "skeleton_result": skeleton_result,
        "graph_result": graph_result,
        "curvature_result": curvature_result,
        "anisotropy_result": anisotropy_result,
        "junction_result": junction_result,
        "kink_result": kink_result,
        "corner_cross_check": corner_cross_check,
    }


def _append(tmp_path, xlsx_path, image_stem, results, gray, region_desc="test region"):
    image_path = Path(f"{image_stem}.jpg")
    out_dir = tmp_path / "outputs"
    out_dir.mkdir(exist_ok=True)
    overlay_path = out_dir / f"{image_stem}_overlay.png"
    rgb = np.stack([gray, gray, gray], axis=-1).astype("uint8")
    render_overlay(
        rgb,
        results["skeleton_result"].skeleton,
        results["graph_result"],
        overlay_path,
        max_overlay_dim=2500,
        junction_result=results["junction_result"],
        kink_result=results["kink_result"],
        corner_cross_check=results["corner_cross_check"],
        detail=False,
    )
    append_chunk_to_workbook(
        xlsx_path,
        image_path=image_path,
        region_desc=region_desc,
        region=None,
        full_image_shape=gray.shape,
        overlay_png_path=overlay_path,
        out_dir=out_dir,
        **results,
    )


def test_header_row_matches_column_headers_constant(tmp_path):
    xlsx_path = tmp_path / "report.xlsx"
    gray = generate_t_junction()
    results = _run_full_pipeline(gray)
    _append(tmp_path, xlsx_path, "image_a", results, gray)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["image_a"]
    header_values = [cell.value for cell in ws[1]]
    assert header_values == COLUMN_HEADERS + IMAGE_COLUMN_HEADERS


def test_data_row_matches_source_dataclasses(tmp_path):
    xlsx_path = tmp_path / "report.xlsx"
    gray = generate_t_junction()
    results = _run_full_pipeline(gray)
    _append(tmp_path, xlsx_path, "image_a", results, gray)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["image_a"]
    row = {header: cell.value for header, cell in zip(COLUMN_HEADERS, ws[2])}

    assert row["otsu_threshold_0to255"] == pytest.approx(results["binarize_result"].threshold)
    assert row["n_edges"] == results["graph_result"].n_edges
    assert row["anisotropy_index_0to1"] == pytest.approx(results["anisotropy_result"].anisotropy_index)
    assert row["n_junctions_T"] == results["junction_result"].n_t
    assert row["n_kinks_flagged"] == len(results["kink_result"].kinks)


def test_second_chunk_of_same_image_appends_a_row(tmp_path):
    xlsx_path = tmp_path / "report.xlsx"
    gray = generate_t_junction()
    results = _run_full_pipeline(gray)

    _append(tmp_path, xlsx_path, "image_a", results, gray, region_desc="chunk 1")
    _append(tmp_path, xlsx_path, "image_a", results, gray, region_desc="chunk 2")

    wb = openpyxl.load_workbook(xlsx_path)
    assert wb.sheetnames == ["image_a"]
    ws = wb["image_a"]
    assert ws.max_row == 3  # header + 2 chunks
    region_col_idx = COLUMN_HEADERS.index("region_description")
    assert ws.cell(row=2, column=region_col_idx + 1).value == "chunk 1"
    assert ws.cell(row=3, column=region_col_idx + 1).value == "chunk 2"


def test_different_image_creates_new_sheet_without_touching_first(tmp_path):
    xlsx_path = tmp_path / "report.xlsx"
    gray = generate_t_junction()
    results = _run_full_pipeline(gray)

    _append(tmp_path, xlsx_path, "image_a", results, gray)
    _append(tmp_path, xlsx_path, "image_b", results, gray)

    wb = openpyxl.load_workbook(xlsx_path)
    assert set(wb.sheetnames) == {"image_a", "image_b"}
    assert wb["image_a"].max_row == 2
    assert wb["image_b"].max_row == 2
    assert "Sheet" not in wb.sheetnames


def test_sanitize_sheet_name_strips_invalid_characters():
    name = _sanitize_sheet_name("weird[name]:with*bad?chars/here\\end", set())
    assert not any(c in name for c in "[]:*?/\\")


def test_sanitize_sheet_name_truncates_long_names():
    long_stem = "a" * 50
    name = _sanitize_sheet_name(long_stem, set())
    assert len(name) <= 31


def test_sanitize_sheet_name_disambiguates_on_truncation_collision():
    stem1 = "x" * 40 + "AAAA"
    stem2 = "x" * 40 + "BBBB"
    name1 = _sanitize_sheet_name(stem1, set())
    name2 = _sanitize_sheet_name(stem2, {name1})
    assert name1 != name2
    assert len(name2) <= 31
